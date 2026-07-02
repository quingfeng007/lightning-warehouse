"""
闪电仓计算工具 - Web 版
基于现有 lightning_warehouse_tool.py 的核心逻辑，包装成 Flask Web 服务。

启动:  python app.py
访问:  http://localhost:5000
"""
import os
import re
import uuid
import shutil
import tempfile
import logging
import time
import json
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from flask import Flask, request, render_template_string, send_file, jsonify, redirect, url_for

# 复用现有工具的核心函数（直接 import 闪电仓工具模块）
# 关键函数: extract_short_name, make_pivot_file2, process_file1
# 为了避免路径问题，这里把核心逻辑重新写一遍（精简版）
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from lightning_warehouse_tool import (
        extract_short_name,
        make_pivot_file2,
        make_pivot_file3,
        make_pivot_file4,
        process_file1,
    )
except ImportError as e:
    # 如果是 PyQt5 缺失错误（云端环境），那是因为 module 顶层 import 了 PyQt5
    # 模块代码已用 try/except 处理，这里再兜底
    if 'PyQt5' not in str(e):
        raise  # 其他 ImportError 才是真问题，重抛
    # 重新尝试，云端环境下 PyQt5 是可选的
    import importlib
    import lightning_warehouse_tool
    extract_short_name = lightning_warehouse_tool.extract_short_name
    make_pivot_file2 = lightning_warehouse_tool.make_pivot_file2
    make_pivot_file3 = lightning_warehouse_tool.make_pivot_file3
    make_pivot_file4 = lightning_warehouse_tool.make_pivot_file4
    process_file1 = lightning_warehouse_tool.process_file1

# ============== Flask 应用 ==============
app = Flask(__name__)
_APP_START_TIME = time.time()
app.secret_key = os.environ.get('FLASK_SECRET_KEY') or os.environ.get('FLASK_SECRET') or os.urandom(32)
APP_PASSWORD = os.environ.get('APP_PASSWORD', 'lightning')
LOCK_TIMEOUT = int(os.environ.get('LOCK_TIMEOUT', '300'))  # 5 分钟锁屏
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB 上限

# 任务存储（用文件存储，以适配 gunicorn 多 worker 场景）
TASK_DIR = os.path.join(tempfile.gettempdir(), 'lightning_tasks')
os.makedirs(TASK_DIR, exist_ok=True)


def _task_path(task_id):
    return os.path.join(TASK_DIR, f'{task_id}.json')


def task_set(task_id, **kwargs):
    """更新任务状态，写入文件"""
    p = _task_path(task_id)
    data = {}
    if os.path.exists(p):
        try:
            with open(p, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception:
            data = {}
    data.update(kwargs)
    data['updated_at'] = time.time()
    # output 是文件路径，不能 json 化，单独存
    output = data.pop('_output', None)
    with open(p, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
    if output:
        with open(p + '.output', 'w', encoding='utf-8') as f:
            f.write(output)


def task_get(task_id):
    """读取任务状态"""
    p = _task_path(task_id)
    if not os.path.exists(p):
        return None
    try:
        with open(p, 'r', encoding='utf-8') as f:
            data = json.load(f)
        out_p = p + '.output'
        if os.path.exists(out_p):
            with open(out_p, 'r', encoding='utf-8') as out_f:
                data['output'] = out_f.read()
        return data
    except Exception:
        return None

# ============== 鉴权 ==============
from functools import wraps
from flask import session, redirect, url_for, request

def require_auth(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get('authed'):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
               'application/json' in request.headers.get('Accept', ''):
                return jsonify(error='未授权,请重新登录', need_login=True), 401
            return redirect(url_for('login', next=request.path))
        # session 过期: 同时检查时间戳
        last = session.get('last_active', 0)
        if time.time() - last > LOCK_TIMEOUT:
            session.clear()
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
               'application/json' in request.headers.get('Accept', ''):
                return jsonify(error='会话已过期,请重新登录', need_login=True), 401
            return redirect(url_for('login', next=request.path, expired=1))
        # 刷新 last_active
        session['last_active'] = time.time()
        return f(*args, **kwargs)
    return wrapped

# ============== 登录页 ==============
LOGIN_HTML = u'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>登录 - 闪电仓计算工具</title>
<style>
body{font-family:-apple-system,"Microsoft YaHei",sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);min-height:100vh;display:flex;align-items:center;justify-content:center;margin:0;padding:20px;}
.box{background:white;padding:40px 36px;border-radius:12px;box-shadow:0 10px 40px rgba(0,0,0,0.2);width:100%;max-width:380px;box-sizing:border-box;}
h1{margin:0 0 8px;color:#2c3e50;font-size:24px;text-align:center;}
p.sub{color:#888;margin:0 0 28px;font-size:13px;text-align:center;}
label{display:block;color:#555;margin-bottom:6px;font-size:14px;}
input[type=password]{width:100%;padding:12px 14px;border:1px solid #ddd;border-radius:6px;font-size:15px;box-sizing:border-box;outline:none;transition:border-color 0.2s;}
input[type=password]:focus{border-color:#667eea;}
button{width:100%;padding:12px;margin-top:18px;background:#667eea;color:white;border:none;border-radius:6px;font-size:16px;font-weight:600;cursor:pointer;transition:background 0.2s;}
button:hover{background:#5568d3;}
button:disabled{background:#999;cursor:not-allowed;}
.err{color:#e74c3c;background:#fde8e8;padding:10px 14px;border-radius:6px;margin-top:14px;font-size:14px;display:none;}
</style>
</head>
<body>
<div class="box">
  <h1>⚡ 闪电仓计算</h1>
  <p class="sub">请输入访问密码</p>
  <form method="post" action="/login" id="loginForm">
    <input type="hidden" name="next" value="{{ next }}">
    <label for="pw">密码</label>
    <input type="password" id="pw" name="password" autofocus required>
    <button type="submit" id="btn">登录</button>
    <div class="err" id="err"></div>
  </form>
</div>
<script>
(function(){
  var params = new URLSearchParams(location.search);
  if (params.get('expired') === '1') {
    var e = document.getElementById('err');
    e.textContent = '会话已过期(5 分钟无操作),请重新登录';
    e.style.display = 'block';
  }
  document.getElementById('loginForm').addEventListener('submit', function(ev){
    ev.preventDefault();
    var btn = document.getElementById('btn');
    btn.disabled = true; btn.textContent = '验证中...';
    var fd = new FormData(this);
    fetch('/login', {method:'POST', body:fd, headers:{'X-Requested-With':'XMLHttpRequest'}})
      .then(function(r){ return r.json(); })
      .then(function(d){
        if (d.ok) {
          location.href = d.next || '/';
        } else {
          var e = document.getElementById('err');
          e.textContent = d.error || '密码错误';
          e.style.display = 'block';
          btn.disabled = false; btn.textContent = '登录';
          document.getElementById('pw').focus();
        }
      })
      .catch(function(){
        var e = document.getElementById('err');
        e.textContent = '网络错误';
        e.style.display = 'block';
        btn.disabled = false; btn.textContent = '登录';
      });
  });
})();
</script>
</body>
</html>'''

# ============== HTML 模板 ==============
INDEX_HTML = u"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>闪电仓计算工具</title>
<style>
body{font-family:-apple-system,"Microsoft YaHei",sans-serif;max-width:760px;margin:30px auto;padding:0 20px;background:#f9f9f9;}
h1{color:#2c3e50;border-bottom:2px solid #3498db;padding-bottom:10px;}
.box{background:white;padding:20px;border-radius:8px;margin:16px 0;box-shadow:0 1px 3px rgba(0,0,0,0.1);}
.row{display:flex;align-items:center;margin:10px 0;gap:10px;}
.row label{width:220px;font-weight:bold;color:#333;}
.row input[type=file]{flex:1;padding:6px;border:1px solid #ddd;border-radius:4px;}
.row .name{width:240px;color:#27ae60;font-size:13px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;}
.cfg label{margin-right:24px;}
.cfg input{padding:6px 10px;border:1px solid #ddd;border-radius:4px;width:80px;}
button{background:#3498db;color:white;border:none;padding:12px 36px;font-size:16px;border-radius:6px;cursor:pointer;margin-top:10px;}
button:hover{background:#2980b9;}
button:disabled{background:#95a5a6;cursor:not-allowed;}
#log{background:#2c3e50;color:#ecf0f1;padding:16px;border-radius:6px;font-family:Menlo,Consolas,monospace;min-height:180px;max-height:400px;overflow-y:auto;white-space:pre-wrap;font-size:13px;line-height:1.5;}
.req{color:#e74c3c;font-weight:bold;}
.opt{color:#95a5a6;font-size:12px;}
#result{display:none;margin-top:16px;padding:16px;background:#d4edda;border-radius:6px;border-left:4px solid #28a745;}
#result a{color:#155724;font-weight:bold;font-size:16px;}
</style>
</head>
<body>
<h1>闪电仓计算工具</h1>
<div class="box">
<form id="uploadForm" enctype="multipart/form-data" method="POST" action="/process">
<div class="row">
<label>1. 闪电仓计算模板 <span class="req">(必填)</span></label>
<input type="file" name="file1" accept=".xlsx,.xls,.xlsm" required>
<span class="name" id="s1"></span>
</div>
<div class="row">
<label>2. 全店数据-门店成交明细 <span class="req">(必填)</span></label>
<input type="file" name="file2" accept=".xlsx,.xls,.xlsm,.csv" required>
<span class="name" id="s2"></span>
</div>
<div class="row">
<label>3. 评价分析明细 <span class="opt">(可选)</span></label>
<input type="file" name="file3" accept=".xlsx,.xls,.xlsm,.csv">
<span class="name" id="s3"></span>
</div>
<div class="row">
<label>4. 门店推广费 <span class="opt">(可选)</span></label>
<input type="file" name="file4" accept=".xlsx,.xls,.xlsm,.csv">
<span class="name" id="s4"></span>
</div>
<div class="cfg" style="background:#f5f5f5;padding:16px;border-radius:8px;margin:20px 0;">
<label>计算天数: <input type="number" name="days" value="30" min="1" max="365" required></label>
<label>体验分参考: <input type="number" name="experience" value="80" min="0" max="100"></label>
</div>
<button type="submit" id="submitBtn">开始处理</button>
</form>
</div>

<h3>处理日志</h3>
<div class="box">
<div id="log">等待开始...</div>
</div>

<div id="result">
<strong>处理完成!</strong> <a id="downloadBtn" href="#" download>下载结果文件</a>
</div>

<script>
(function(){
  function $(id){return document.getElementById(id);}

  // 选中文件后显示文件名
  var files=document.querySelectorAll('input[type=file]');
  for(var i=0;i<files.length;i++){
    (function(inp,idx){
      inp.addEventListener('change',function(){
        var f=inp.files[0];
        $('s'+(idx+1)).textContent=f?('已选择: '+f.name):'';
      });
    })(files[i],i);
  }

  var form=$('uploadForm');
  var btn=$('submitBtn');
  var log=$('log');
  var result=$('result');

  form.addEventListener('submit',function(e){
    e.preventDefault();
    btn.disabled=true;
    btn.textContent='处理中...';
    log.textContent='正在上传文件,请稍候...';
    result.style.display='none';

    var fd=new FormData(form);

    fetch('/process',{method:'POST',body:fd,headers:{'X-Requested-With':'XMLHttpRequest','Accept':'application/json'}})
      .then(function(r){return r.json();})
      .then(function(data){
        if(data.task_id){
          pollLog(data.task_id);
        }else{
          log.textContent='错误: '+(data.error||'未知错误');
          btn.disabled=false;
          btn.textContent='开始处理';
        }
      })
      .catch(function(err){
        log.textContent='网络错误: '+err.message;
        btn.disabled=false;
        btn.textContent='开始处理';
      });
  });

  function pollLog(tid){
    var t=setInterval(function(){
      fetch('/status/'+tid)
        .then(function(r){return r.json();})
        .then(function(d){
          log.textContent=d.log||'';
          log.scrollTop=log.scrollHeight;
          if(d.status==='done'){
            clearInterval(t);
            btn.disabled=false;
            btn.textContent='开始处理';
            result.style.display='block';
            $('downloadBtn').href='/download/'+tid;
          }else if(d.status==='error'){
            clearInterval(t);
            btn.disabled=false;
            btn.textContent='开始处理';
            log.textContent=log.textContent+'\n\n[错误] '+d.error;
          }
        })
        .catch(function(e){});
    },1000);
  }
})();
</script>

<script>
// 5 分钟锁屏(不刷新也计时,刷新也重新计时)
(function(){
  var TIMEOUT_MS = 5 * 60 * 1000;
  var warnTimer = null, lockTimer = null, heartbeatTimer = null, warned = false;

  function lock(){
    fetch('/logout', {method:'POST', headers:{'X-Requested-With':'XMLHttpRequest'}})
      .finally(function(){ location.href = '/login?expired=1'; });
  }

  function showWarn(){
    if (document.getElementById('lockOverlay')) return;
    var o = document.createElement('div');
    o.id = 'lockOverlay';
    o.style.cssText = 'position:fixed;top:0;left:0;right:0;background:#f39c12;color:white;padding:14px;text-align:center;z-index:99999;font-size:15px;box-shadow:0 2px 6px rgba(0,0,0,0.2);';
    o.textContent = '\u26a0\ufe0f 30 秒后将自动锁定(5 分钟无操作),点任意位置可继续使用';
    document.body.appendChild(o);
  }

  function reset(){
    var o = document.getElementById('lockOverlay');
    if (o) o.remove();
    warned = false;
    clearTimeout(warnTimer);
    clearTimeout(lockTimer);
    warnTimer = setTimeout(function(){ warned = true; showWarn(); }, TIMEOUT_MS - 30000);
    lockTimer = setTimeout(lock, TIMEOUT_MS);
  }

  // 心跳:每 1 分钟 ping 一次,服务端 session 计时同步
  setInterval(function(){
    fetch('/heartbeat', {method:'POST', headers:{'X-Requested-With':'XMLHttpRequest'}})
      .then(function(r){ if (r.status === 401) lock(); return r.json(); })
      .catch(function(){});
  }, 60000);

  ['mousemove','mousedown','keydown','scroll','touchstart','click'].forEach(function(ev){
    document.addEventListener(ev, reset, {passive:true});
  });
  document.addEventListener('visibilitychange', function(){
    if (!document.hidden) reset();
  });

  reset();
})();
</script>
</body>
</html>
"""

# ============== 路由 ==============

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        pw = request.form.get('password', '')
        next_url = request.form.get('next') or '/'
        if pw == APP_PASSWORD:
            session['authed'] = True
            session['last_active'] = time.time()
            session.permanent = False
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            if is_ajax:
                return jsonify(ok=True, next=next_url)
            return redirect(next_url)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        if is_ajax:
            return jsonify(ok=False, error='密码错误')
        return render_template_string(LOGIN_HTML, next=next_url), 401
    # GET
    return render_template_string(LOGIN_HTML, next=request.args.get('next', '/'))

@app.route('/logout', methods=['POST', 'GET'])
def logout():
    session.clear()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
       'application/json' in request.headers.get('Accept', ''):
        return jsonify(ok=True)
    return redirect(url_for('login'))

# 心跳: 客户端定时调用，刷新 last_active
@app.route('/heartbeat', methods=['POST'])
@require_auth
def heartbeat():
    return jsonify(ok=True, last_active=session.get('last_active'))

@app.route('/')
@require_auth
def index():
    return render_template_string(INDEX_HTML)

@app.route('/health')
def health():
    return jsonify(
        status='ok',
        version='1.0',
        uptime=time.time() - _APP_START_TIME,
        timestamp=time.time(),
    )

# 极轻量 ping 端点 — 供 cron-job.org 保持 Render 实例唤醒
# 返回纯文本 "pong",状态码 200,耗时 < 5ms
@app.route('/ping')
def ping():
    return 'pong', 200, {'Content-Type': 'text/plain; charset=utf-8', 'Cache-Control': 'no-store'}

RESULT_HTML = u"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>处理结果 - 闪电仓计算工具</title>
<style>
body{font-family:-apple-system,"Microsoft YaHei",sans-serif;max-width:760px;margin:30px auto;padding:0 20px;background:#f9f9f9;}
h1{color:#2c3e50;border-bottom:2px solid #3498db;padding-bottom:10px;}
.box{background:white;padding:20px;border-radius:8px;margin:16px 0;box-shadow:0 1px 3px rgba(0,0,0,0.1);}
#log{background:#2c3e50;color:#ecf0f1;padding:16px;border-radius:6px;font-family:Menlo,Consolas,monospace;min-height:300px;max-height:500px;overflow-y:auto;white-space:pre-wrap;font-size:13px;line-height:1.5;}
#result{display:none;margin-top:16px;padding:20px;background:#d4edda;border-radius:6px;border-left:4px solid #28a745;text-align:center;}
#result a{color:#155724;font-weight:bold;font-size:18px;display:inline-block;margin-top:10px;padding:12px 24px;background:#28a745;color:white;text-decoration:none;border-radius:6px;}
#result a:hover{background:#218838;}
.btn-home{display:inline-block;margin-top:10px;padding:8px 16px;background:#6c757d;color:white;text-decoration:none;border-radius:4px;font-size:14px;}
.btn-home:hover{background:#5a6268;}
.status{padding:12px;border-radius:6px;margin:10px 0;font-weight:bold;}
.status.running{background:#fff3cd;color:#856404;border-left:4px solid #ffc107;}
.status.done{background:#d4edda;color:#155724;border-left:4px solid #28a745;}
.status.error{background:#f8d7da;color:#721c24;border-left:4px solid #dc3545;}
</style>
</head>
<body>
<h1>闪电仓计算 - 处理结果</h1>
<div class="box">
<div id="statusBox" class="status running">处理中...</div>
<h3>处理日志</h3>
<div id="log">初始化...</div>
<div id="result">
<strong>处理完成!</strong><br>
<a id="downloadBtn" href="#" download>下载结果文件</a>
<br>
<a class="btn-home" href="/">返回处理新文件</a>
</div>
</div>

<script>
(function(){
  var taskId='__TASK_ID__';
  var log=document.getElementById('log');
  var statusBox=document.getElementById('statusBox');
  var result=document.getElementById('result');

  var t=setInterval(function(){
    fetch('/status/'+taskId)
      .then(function(r){return r.json();})
      .then(function(d){
        log.textContent=d.log||'';
        log.scrollTop=log.scrollHeight;
        if(d.status==='done'){
          clearInterval(t);
          statusBox.className='status done';
          statusBox.textContent='处理完成!';
          result.style.display='block';
          document.getElementById('downloadBtn').href='/download/'+taskId;
        }else if(d.status==='error'){
          clearInterval(t);
          statusBox.className='status error';
          statusBox.textContent='处理失败';
          log.textContent=log.textContent+'\\n\\n[错误] '+d.error;
        }
      })
      .catch(function(e){
        log.textContent=log.textContent+'\\n[网络错误] '+e.message;
      });
  },1000);
})();
</script>

<script>
// 5 分钟锁屏(同主页逻辑)
(function(){
  var TIMEOUT_MS = 5 * 60 * 1000;
  var warnTimer = null, lockTimer = null, warned = false;

  function lock(){
    fetch('/logout', {method:'POST', headers:{'X-Requested-With':'XMLHttpRequest'}})
      .finally(function(){ location.href = '/login?expired=1'; });
  }

  function showWarn(){
    if (document.getElementById('lockOverlay')) return;
    var o = document.createElement('div');
    o.id = 'lockOverlay';
    o.style.cssText = 'position:fixed;top:0;left:0;right:0;background:#f39c12;color:white;padding:14px;text-align:center;z-index:99999;font-size:15px;box-shadow:0 2px 6px rgba(0,0,0,0.2);';
    o.textContent = '\u26a0\ufe0f 30 秒后将自动锁定,点任意位置可继续使用';
    document.body.appendChild(o);
  }

  function reset(){
    var o = document.getElementById('lockOverlay');
    if (o) o.remove();
    warned = false;
    clearTimeout(warnTimer);
    clearTimeout(lockTimer);
    warnTimer = setTimeout(function(){ warned = true; showWarn(); }, TIMEOUT_MS - 30000);
    lockTimer = setTimeout(lock, TIMEOUT_MS);
  }

  setInterval(function(){
    fetch('/heartbeat', {method:'POST', headers:{'X-Requested-With':'XMLHttpRequest'}})
      .then(function(r){ if (r.status === 401) lock(); return r.json(); })
      .catch(function(){});
  }, 60000);

  ['mousemove','mousedown','keydown','scroll','touchstart','click'].forEach(function(ev){
    document.addEventListener(ev, reset, {passive:true});
  });
  document.addEventListener('visibilitychange', function(){
    if (!document.hidden) reset();
  });

  reset();
})();
</script>
</body>
</html>
"""


@app.route('/process', methods=['POST'])
@require_auth
def process():
    """接收文件，启动后台处理。
    - 如果是 fetch/AJAX(X-Requested-With 或 Accept: application/json),返回 JSON
    - 如果是浏览器原生表单提交,返回 HTML 页面(含自动轮询 JS)
    """
    # 检查必填
    is_ajax_check = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    if 'file1' not in request.files:
        if is_ajax_check:
            return jsonify(error='缺少文件: file1'), 400
        return '<h1>错误</h1><p>缺少文件 file1</p>', 400
    if 'file2' not in request.files:
        if is_ajax_check:
            return jsonify(error='缺少文件: file2'), 400
        return '<h1>错误</h1><p>缺少文件 file2</p>', 400
    if 'days' not in request.form:
        if is_ajax_check:
            return jsonify(error='缺少计算天数'), 400
        return '<h1>错误</h1><p>缺少计算天数</p>', 400

    days = int(request.form['days'])
    experience = int(request.form.get('experience', 80))

    # 保存文件
    task_id = uuid.uuid4().hex[:12]
    work_dir = os.path.join(tempfile.gettempdir(), f'lightning_{task_id}')
    os.makedirs(work_dir, exist_ok=True)

    file1_path = os.path.join(work_dir, f'file1_{request.files["file1"].filename}')
    file2_path = os.path.join(work_dir, f'file2_{request.files["file2"].filename}')
    file1 = request.files['file1']
    file2 = request.files['file2']
    file1.save(file1_path)
    file2.save(file2_path)

    file3_path = file4_path = None
    if 'file3' in request.files and request.files['file3'].filename:
        file3_path = os.path.join(work_dir, f'file3_{request.files["file3"].filename}')
        request.files['file3'].save(file3_path)
    if 'file4' in request.files and request.files['file4'].filename:
        file4_path = os.path.join(work_dir, f'file4_{request.files["file4"].filename}')
        request.files['file4'].save(file4_path)

    # 初始化任务
    task_set(task_id, status='running', log='', work_dir=work_dir)

    # 判断客户端类型
    # 浏览器原生表单提交:无 X-Requested-With,Accept 为 */* 或 text/html
    # fetch 提交:带 X-Requested-With: XMLHttpRequest
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if is_ajax:
        # fetch 路径:返回 JSON task_id,让前端轮询
        # 把后台启动逻辑延后到 jsonify 之后(只让 ajax 路径走)
        import threading
        def run():
            try:
                log_lines = []
                def log(msg):
                    log_lines.append(msg)
                    task_set(task_id, log='\n'.join(log_lines))

                log(f'[1/4] 处理文件2: 全店数据-门店成交明细')
                tmp2 = os.path.join(work_dir, '_pivot_file2.xlsx')
                df2_stores = []
                wb1 = load_workbook(file1_path, data_only=False)
                ws1 = wb1.active
                for row in ws1.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                    v = row[0]
                    if v is not None and str(v).strip():
                        df2_stores.append(str(v).strip())
                log(f'  门店数: {len(df2_stores)}')
                make_pivot_file2(file2_path, df2_stores, tmp2, log)
                log(f'  透视完成: {tmp2}')

                tmp3 = None
                if file3_path:
                    log(f'[2/4] 处理文件3: 评价分析明细')
                    tmp3 = os.path.join(work_dir, '_pivot_file3.xlsx')
                    make_pivot_file3(file3_path, df2_stores, tmp3, log)
                    log(f'  透视完成: {tmp3}')

                tmp4 = None
                if file4_path:
                    log(f'[3/4] 处理文件4: 门店推广费')
                    tmp4 = os.path.join(work_dir, '_pivot_file4.xlsx')
                    make_pivot_file4(file4_path, df2_stores, tmp4, log)
                    log(f'  透视完成: {tmp4}')

                log(f'[4/4] 生成最终结果到文件1')
                output_path = os.path.join(work_dir, '闪电仓计算结果.xlsx')
                process_file1(file1_path, days, tmp2, output_path, log,
                              file3_vlookup_path=tmp3,
                              file4_vlookup_path=tmp4)

                log(f'Done: {output_path}')
                task_set(task_id, status='done', _output=output_path)
            except Exception as e:
                import traceback
                err_msg = str(e) + '\n' + traceback.format_exc()
                log_lines.append(f'\nERROR: {err_msg}')
                task_set(task_id, status='error', error=err_msg, log='\n'.join(log_lines))
        threading.Thread(target=run, daemon=True).start()
        return jsonify(task_id=task_id)

    # 浏览器原生表单提交路径:启动后台处理 + 返回 HTML 页面(含自动轮询)
    import threading
    def run_html():
        try:
            log_lines = []
            def log(msg):
                log_lines.append(msg)
                task_set(task_id, log='\n'.join(log_lines))

            log(f'[1/4] 处理文件2: 全店数据-门店成交明细')
            tmp2 = os.path.join(work_dir, '_pivot_file2.xlsx')
            df2_stores = []
            wb1 = load_workbook(file1_path, data_only=False)
            ws1 = wb1.active
            for row in ws1.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                v = row[0]
                if v is not None and str(v).strip():
                    df2_stores.append(str(v).strip())
            log(f'  门店数: {len(df2_stores)}')
            make_pivot_file2(file2_path, df2_stores, tmp2, log)
            log(f'  透视完成: {tmp2}')

            tmp3 = None
            if file3_path:
                log(f'[2/4] 处理文件3: 评价分析明细')
                tmp3 = os.path.join(work_dir, '_pivot_file3.xlsx')
                make_pivot_file3(file3_path, df2_stores, tmp3, log)
                log(f'  透视完成: {tmp3}')

            tmp4 = None
            if file4_path:
                log(f'[3/4] 处理文件4: 门店推广费')
                tmp4 = os.path.join(work_dir, '_pivot_file4.xlsx')
                make_pivot_file4(file4_path, df2_stores, tmp4, log)
                log(f'  透视完成: {tmp4}')

            log(f'[4/4] 生成最终结果到文件1')
            output_path = os.path.join(work_dir, '闪电仓计算结果.xlsx')
            process_file1(file1_path, days, tmp2, output_path, log,
                          file3_vlookup_path=tmp3,
                          file4_vlookup_path=tmp4)

            log(f'Done: {output_path}')
            task_set(task_id, status='done', _output=output_path)
        except Exception as e:
            import traceback
            err_msg = str(e) + '\n' + traceback.format_exc()
            log_lines.append(f'\nERROR: {err_msg}')
            task_set(task_id, status='error', error=err_msg, log='\n'.join(log_lines))
    threading.Thread(target=run_html, daemon=True).start()
    return RESULT_HTML.replace('__TASK_ID__', task_id)


@app.route('/status/<task_id>')
@require_auth
def status(task_id):
    t = task_get(task_id)
    if not t:
        return jsonify(error='任务不存在'), 404
    return jsonify(status=t.get('status', 'running'), log=t.get('log', ''), error=t.get('error'))

@app.route('/download/<task_id>')
@require_auth
def download(task_id):
    t = task_get(task_id)
    if not t or t.get('status') != 'done':
        return jsonify(error='任务未完成'), 400
    out = t.get('output')
    if not out or not os.path.exists(out):
        return jsonify(error='结果文件丢失'), 404
    return send_file(out, as_attachment=True, download_name='result.xlsx')

# ============== 启动 ==============
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
