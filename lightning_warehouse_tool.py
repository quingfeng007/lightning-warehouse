# -*- coding: utf-8 -*-
"""
闪电仓计算工具 v1.0
- 4个文件上传（先实现文件1+文件2的核心功能）
- 自动透视、简称提取、VLOOKUP
- 后续文件3/4留接口
"""

import sys
import os
import re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QLineEdit, QFileDialog,
    QTextEdit, QMessageBox, QGroupBox, QGridLayout
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QFont


# ============ 核心处理逻辑 ============

def extract_short_name(full_name: str) -> str:
    """从"鲜果派对·甘草水果（XX店）"中提取括号里的店名简称"""
    if not full_name or not isinstance(full_name, str):
        return ""
    m = re.search(r"[\(（]([^\)）]+)[\)）]", full_name)
    if m:
        return m.group(1).strip()
    return full_name.strip()


def _read_table_auto(path: str, log_func=None) -> pd.DataFrame:
    """
    根据后缀自动选择读取方式，支持 .xlsx / .xls / .xlsm / .csv
    返回的是无 header 的 DataFrame（跟原 read_excel header=None 一致）
    """
    log = log_func or (lambda msg: None)
    ext = os.path.splitext(path)[1].lower()
    log(f"[读取] 后缀: {ext}")

    if ext in (".xlsx", ".xlsm", ".xls"):
        return pd.read_excel(path, sheet_name=0, header=None)
    elif ext == ".csv":
        # 尝试多种编码（GBK / UTF-8 / UTF-8-BOM）
        # 不强制 dtype=str，让 pandas 自动推断数字类型
        df = None
        for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
            try:
                df = pd.read_csv(path, header=None, encoding=enc).fillna("")
                break
            except UnicodeDecodeError:
                continue
        if df is None:
            # 兜底
            df = pd.read_csv(path, header=None, encoding="gbk", encoding_errors="ignore").fillna("")

        # 不在 _read_table_auto 内部转换数字列——如果表头行被转成 NaN 会影响上层识别列名
        # 转换逻辑统一在 make_pivot_file2 里、剥离表头后进行
        return df
    else:
        raise ValueError(f"不支持的文件格式: {ext}（仅支持 .xlsx/.xls/.xlsm/.csv）")


def make_pivot_file2(file2_path: str, file1_stores: list, output_path: str, log_func=None):
    """
    处理文件2：全店数据-门店成交明细
    - 读取原始数据（支持 xlsx/xls/xlsm/csv）
    - 对 D/L/M/P/V 列做透视（行=商家名称，值=4列求和）
    - H列右侧插入I列（简称）
    - N列贴门店名，O-R列写VLOOKUP公式（先指向J-M）
    """
    log = log_func or (lambda msg: None)
    log(f"[文件2] 读取: {file2_path}")

    # 用 pandas 读原始数据（自动识别后缀）
    df = _read_table_auto(file2_path, log)
    log(f"[文件2] 原始数据形状: {df.shape}")

    # 找标题行（"商家名称"所在行）
    header_row = None
    store_col_idx = None
    for i in range(min(10, len(df))):
        for j in range(min(20, df.shape[1])):
            v = df.iat[i, j]
            if isinstance(v, str) and "商家名称" in v:
                header_row = i
                store_col_idx = j
                break
        if header_row is not None:
            break

    if header_row is None or store_col_idx is None:
        raise ValueError("文件2中找不到'商家名称'标题行，请检查文件格式")

    log(f"[文件2] 标题行: 第 {header_row + 1} 行, 商家名称列: 第 {store_col_idx + 1} 列")

    # 把标题行设为列名
    df.columns = df.iloc[header_row].tolist()
    # 跳过标题行
    df = df.iloc[header_row + 1:].reset_index(drop=True)

    # 强制把可能的数值列转为数字（避免 groupby 字符串拼接）
    for col in df.columns:
        if df[col].dtype == object or pd.api.types.is_string_dtype(df[col]):
            converted = pd.to_numeric(df[col], errors="coerce")
            non_na_ratio = converted.notna().sum() / max(len(converted), 1)
            if non_na_ratio > 0.5:
                df[col] = converted

    # 按列名取关键列（不依赖列位置，避免实际文件列顺序不同）
    store_col = df.columns[store_col_idx]  # 用于 groupby

    def find_col(df, *keywords):
        """根据关键字查找匹配的列名（不区分全半角括号）"""
        for c in df.columns:
            cn = str(c).replace(" ", "").replace("（", "(").replace("）", ")")
            for kw in keywords:
                kwn = kw.replace(" ", "").replace("（", "(").replace("）", ")")
                if kwn in cn:
                    return c
        return None

    col_L = find_col(df, "(日均)在线营业时长", "在线营业时长")  # (日均)在线营业时长(h)
    col_M = find_col(df, "推单数")  # 推单数
    col_P = find_col(df, "实付交易额(推单日期)", "实付交易额")  # 实付交易额
    col_V = find_col(df, "缺货导致退款订单数", "缺货退款订单数")  # 缺货导致退款订单数

    # 找不到时报错
    missing = []
    for name, col in [("L在线时长", col_L), ("M推单数", col_M), ("P实付交易额", col_P), ("V缺货退款", col_V)]:
        if col is None:
            missing.append(name)
    if missing:
        raise ValueError(f"文件2 找不到关键列: {', '.join(missing)}\n现有列名: {list(df.columns)}")

    log(f"[文件2] 商家名称列: {store_col!r} (idx={store_col_idx})")
    log(f"[文件2] 总行数: {len(df)}")

    # 透视：按商家名称分组，对 D(计数) / L/M/P/V(求和)聚合
    pivot_count = df.groupby(store_col).size().reset_index(name="计数_D")
    pivot_l = df.groupby(store_col)[col_L].sum().reset_index(name="在线时长_L")
    pivot_m = df.groupby(store_col)[col_M].sum().reset_index(name="推单数_M")
    pivot_p = df.groupby(store_col)[col_P].sum().reset_index(name="实付交易额_P")
    pivot_v = df.groupby(store_col)[col_V].sum().reset_index(name="缺货退款_V")

    # 合并
    pivot = pivot_count.merge(pivot_l, on=store_col, how="outer") \
                       .merge(pivot_m, on=store_col, how="outer") \
                       .merge(pivot_p, on=store_col, how="outer") \
                       .merge(pivot_v, on=store_col, how="outer") \
                       .fillna(0)

    log(f"[文件2] 透视后商家数: {len(pivot)}")

    # 用 openpyxl 写新文件，保留公式
    wb = Workbook()
    ws = wb.active
    ws.title = "透视结果"

    # 表头（H-M）
    headers = [
        ("H", "商家名称"),
        ("I", "简称"),
        ("J", "在线时长"),
        ("K", "推单数"),
        ("L", "实付交易额"),
        ("M", "缺货退款订单数"),
    ]
    for col, h in headers:
        ws[f"{col}1"] = h
        ws[f"{col}1"].font = Font(bold=True)
        ws[f"{col}1"].fill = PatternFill("solid", fgColor="DDEBF7")

    # 写入透视数据（H2 起到 H{1+n}）
    for i, row in pivot.iterrows():
        r = i + 2
        ws[f"H{r}"] = row[store_col]
        ws[f"I{r}"] = extract_short_name(row[store_col])
        ws[f"J{r}"] = float(row["在线时长_L"])
        ws[f"K{r}"] = float(row["推单数_M"])
        ws[f"L{r}"] = float(row["实付交易额_P"])
        ws[f"M{r}"] = float(row["缺货退款_V"])

    # N列：门店，贴文件1的A列门店名
    ws["N1"] = "门店"
    ws["N1"].font = Font(bold=True)
    ws["N1"].fill = PatternFill("solid", fgColor="DDEBF7")
    for i, store in enumerate(file1_stores):
        r = i + 2
        ws[f"N{r}"] = store

    # O-R列：VLOOKUP 公式
    # 查找值=N列门店, 范围=$I$2:$M$N, 返回列分别=2(J在线时长)/3(K推单数)/4(L实付交易额)/5(M缺货退款)
    vlookup_cols = [
        ("O", 2, "在线时长"),
        ("P", 3, "推单数"),
        ("Q", 4, "实付交易额"),
        ("R", 5, "缺货退款订单数"),
    ]
    for col, idx, _ in vlookup_cols:
        ws[f"{col}1"] = f"匹配_{_}"
        ws[f"{col}1"].font = Font(bold=True)
        ws[f"{col}1"].fill = PatternFill("solid", fgColor="DDEBF7")

    last_pivot_row = len(pivot) + 1
    # 用 Python 做 VLOOKUP（不写公式，直接填值）—— 按 N列门店匹配 I列简称
    # 构建查找表：{简称: (在线时长, 推单数, 实付交易额, 缺货退款订单数)}
    lookup = {}
    for _, row in pivot.iterrows():
        key = extract_short_name(row[store_col])
        lookup[key] = (
            float(row["在线时长_L"]),
            float(row["推单数_M"]),
            float(row["实付交易额_P"]),
            float(row["缺货退款_V"]),
        )
    vlookup_value_cols = [
        ("O", 0, "在线时长"),
        ("P", 1, "推单数"),
        ("Q", 2, "实付交易额"),
        ("R", 3, "缺货退款订单数"),
    ]
    for i, store in enumerate(file1_stores):
        r = i + 2
        matched = lookup.get(store, (0.0, 0.0, 0.0, 0.0))
        for col, val_idx, _ in vlookup_value_cols:
            ws[f"{col}{r}"] = matched[val_idx]

    # 列宽
    ws.column_dimensions["H"].width = 35
    ws.column_dimensions["I"].width = 15
    for c in ["J", "K", "L", "M", "N", "O", "P", "Q", "R"]:
        ws.column_dimensions[c].width = 14

    wb.save(output_path)
    log(f"[文件2] 已保存: {output_path}")
    return output_path


def make_pivot_file3(file3_path: str, file1_stores: list, output_path: str, log_func=None):
    """
    处理文件3：评价分析明细
    - Sheet1 中差评数：D列=店铺名称, N列=商家评分(计数) 筛 1,2,3
    - Sheet2 评分平均：D列=店铺名称, N列=商家评分(平均值) 筛 1,2,3,4,5
    - 布局：D=店铺名称, E=简称, F=值, G=门店, H=匹配
    """
    log = log_func or (lambda msg: None)
    log(f"[文件3] 读取: {file3_path}")

    df = _read_table_auto(file3_path, log)
    log(f"[文件3] 原始数据形状: {df.shape}")

    # 找标题行（"店铺名称"或"商家名称"所在行）
    header_row = None
    store_col_idx = None
    for i in range(min(10, len(df))):
        for j in range(min(20, df.shape[1])):
            v = df.iat[i, j]
            if isinstance(v, str) and ("店铺名称" in v or "商家名称" in v):
                header_row = i
                store_col_idx = j
                break
        if header_row is not None:
            break
    if header_row is None:
        raise ValueError("文件3中找不到'店铺名称'/'商家名称'标题行")
    log(f"[文件3] 标题行: 第 {header_row + 1} 行, 店铺名称列: 第 {store_col_idx + 1} 列")

    # 找"商家评分"列
    raw_df = _read_table_auto(file3_path, log)
    score_col_idx = None
    score_col_name = None
    for j in range(raw_df.shape[1]):
        v = raw_df.iat[header_row, j] if header_row < len(raw_df) else None
        if isinstance(v, str) and ("评分" in v):
            score_col_idx = j
            score_col_name = v
            break
    if score_col_idx is None:
        raise ValueError("文件3中找不到'商家评分'列")
    log(f"[文件3] 商家评分列: idx={score_col_idx} name={score_col_name}")

    df = df.iloc[header_row + 1:].reset_index(drop=True)

    # 转换数值列
    for col in df.columns:
        if df[col].dtype == object or pd.api.types.is_string_dtype(df[col]):
            converted = pd.to_numeric(df[col], errors="coerce")
            if converted.notna().sum() / max(len(converted), 1) > 0.5:
                df[col] = converted

    store_col = df.columns[store_col_idx]
    score_series = df.iloc[:, score_col_idx]
    log(f"[文件3] 店铺名称列: {store_col}, 总行数: {len(df)}")

    # ===== Sheet1: 中差评数（计数, 筛 1,2,3）=====
    mask1 = score_series.isin([1, 2, 3, 1.0, 2.0, 3.0])
    df1 = df[mask1]
    pivot1 = df1.groupby(store_col).size().reset_index(name="中差评数")
    log(f"[文件3] Sheet1(中差评数) 门店数: {len(pivot1)}")

    # ===== Sheet2: 评分平均（平均值, 筛 1,2,3,4,5）=====
    mask2 = score_series.isin([1, 2, 3, 4, 5, 1.0, 2.0, 3.0, 4.0, 5.0])
    df2 = df[mask2]
    pivot2 = df2.groupby(store_col)[df2.columns[score_col_idx]].mean().reset_index(name="评分")
    log(f"[文件3] Sheet2(评分平均) 门店数: {len(pivot2)}")

    # 写 xlsx
    wb = Workbook()

    # --- Sheet1: 中差评数 ---
    ws1 = wb.active
    ws1.title = "中差评数"
    headers1 = [
        ("D", "店铺名称"),
        ("E", "简称"),
        ("F", "中差评数"),
        ("G", "门店"),
        ("H", "匹配_中差评数"),
    ]
    for col, h in headers1:
        ws1[f"{col}1"] = h
        ws1[f"{col}1"].font = Font(bold=True)
        ws1[f"{col}1"].fill = PatternFill("solid", fgColor="DDEBF7")
    for i, row in pivot1.iterrows():
        r = i + 2
        ws1[f"D{r}"] = row[store_col]
        ws1[f"E{r}"] = extract_short_name(row[store_col])
        ws1[f"F{r}"] = int(row["中差评数"])
    for i, store in enumerate(file1_stores):
        r = i + 2
        ws1[f"G{r}"] = store
    lookup1 = {}
    for _, row in pivot1.iterrows():
        key = extract_short_name(row[store_col])
        lookup1[key] = int(row["中差评数"])
    for i, store in enumerate(file1_stores):
        r = i + 2
        ws1[f"H{r}"] = lookup1.get(store, 0)

    # --- Sheet2: 评分平均 ---
    ws2 = wb.create_sheet("评分")
    headers2 = [
        ("D", "店铺名称"),
        ("E", "简称"),
        ("F", "评分"),
        ("G", "门店"),
        ("H", "匹配_评分"),
    ]
    for col, h in headers2:
        ws2[f"{col}1"] = h
        ws2[f"{col}1"].font = Font(bold=True)
        ws2[f"{col}1"].fill = PatternFill("solid", fgColor="DDEBF7")
    for i, row in pivot2.iterrows():
        r = i + 2
        ws2[f"D{r}"] = row[store_col]
        ws2[f"E{r}"] = extract_short_name(row[store_col])
        ws2[f"F{r}"] = round(float(row["评分"]), 4)
    for i, store in enumerate(file1_stores):
        r = i + 2
        ws2[f"G{r}"] = store
    lookup2 = {}
    for _, row in pivot2.iterrows():
        key = extract_short_name(row[store_col])
        lookup2[key] = round(float(row["评分"]), 4)
    for i, store in enumerate(file1_stores):
        r = i + 2
        ws2[f"H{r}"] = lookup2.get(store, 0.0)

    # 列宽
    for ws_ in (ws1, ws2):
        ws_.column_dimensions["D"].width = 35
        ws_.column_dimensions["E"].width = 15
        for c in ["F", "G", "H"]:
            ws_.column_dimensions[c].width = 16

    wb.save(output_path)
    log(f"[文件3] 已保存: {output_path}")
    return output_path


def make_pivot_file4(file4_path: str, file1_stores: list, output_path: str, log_func=None):
    """
    处理文件4：门店推广费
    - 透视：D列=门店名称(计数), M列=推广消费实付(元)(求和)
    - 输出布局：D=门店名称, E=简称, F=推广费, G=门店, H=匹配_推广费
    """
    log = log_func or (lambda msg: None)
    log(f"[文件4] 读取: {file4_path}")

    df = _read_table_auto(file4_path, log)
    log(f"[文件4] 原始数据形状: {df.shape}")

    # 找标题行（"门店名称"所在行）
    header_row = None
    store_col_idx = None
    for i in range(min(10, len(df))):
        for j in range(min(20, df.shape[1])):
            v = df.iat[i, j]
            if isinstance(v, str) and ("门店名称" in v or "店铺名称" in v):
                header_row = i
                store_col_idx = j
                break
        if header_row is not None:
            break
    if header_row is None:
        raise ValueError("文件4中找不到'门店名称'/'店铺名称'标题行")
    log(f"[文件4] 标题行: 第 {header_row + 1} 行, 门店名称列: 第 {store_col_idx + 1} 列")

    df = df.iloc[header_row + 1:].reset_index(drop=True)

    # 转换数值列
    for col in df.columns:
        if df[col].dtype == object or pd.api.types.is_string_dtype(df[col]):
            converted = pd.to_numeric(df[col], errors="coerce")
            if converted.notna().sum() / max(len(converted), 1) > 0.5:
                df[col] = converted

    # 找"推广消费实付"列（直接查标题行的值，不依赖 df.columns）
    cost_col_idx = None
    cost_col_name = None
    raw_df = _read_table_auto(file4_path, log)
    for j in range(raw_df.shape[1]):
        v = raw_df.iat[header_row, j] if header_row < len(raw_df) else None
        if isinstance(v, str) and ("推广消费" in v or "推广费" in v or "消费实付" in v):
            cost_col_idx = j
            cost_col_name = v
            break
    if cost_col_idx is None:
        raise ValueError("文件4中找不到'推广消费实付'列")
    log(f"[文件4] 推广费列: idx={cost_col_idx} name={cost_col_name}")

    store_col = df.columns[store_col_idx]
    log(f"[文件4] 门店名称列: {store_col}")

    # 透视：门店名称分组，推广费求和
    pivot = df.groupby(store_col)[df.columns[cost_col_idx]].sum().reset_index(name="推广费")
    log(f"[文件4] 透视后门店数: {len(pivot)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "透视结果"

    # 表头 D~H
    headers = [
        ("D", "门店名称"),
        ("E", "简称"),
        ("F", "推广费"),
        ("G", "门店"),
        ("H", "匹配_推广费"),
    ]
    for col, h in headers:
        ws[f"{col}1"] = h
        ws[f"{col}1"].font = Font(bold=True)
        ws[f"{col}1"].fill = PatternFill("solid", fgColor="DDEBF7")

    # 写入透视数据
    for i, row in pivot.iterrows():
        r = i + 2
        ws[f"D{r}"] = row[store_col]
        ws[f"E{r}"] = extract_short_name(row[store_col])
        ws[f"F{r}"] = float(row["推广费"])

    # G 列：贴门店名
    for i, store in enumerate(file1_stores):
        r = i + 2
        ws[f"G{r}"] = store

    # 用 Python 做 VLOOKUP：H = F列值（按 G门店匹配 E简称）
    lookup = {}
    for _, row in pivot.iterrows():
        key = extract_short_name(row[store_col])
        lookup[key] = float(row["推广费"])
    for i, store in enumerate(file1_stores):
        r = i + 2
        ws[f"H{r}"] = lookup.get(store, 0.0)

    # 列宽
    ws.column_dimensions["D"].width = 35
    ws.column_dimensions["E"].width = 15
    for c in ["F", "G", "H"]:
        ws.column_dimensions[c].width = 16

    wb.save(output_path)
    log(f"[文件4] 已保存: {output_path}")
    return output_path


def process_file1(file1_path: str, days: int, file2_vlookup_path: str, output_path: str,
                   log_func=None, file3_vlookup_path: str = None, file4_vlookup_path: str = None,
                   product_ref: float = 0.9, experience_score: float = 80.0):
    """
    处理文件1：闪电仓计算模板
    - 读取A列门店名
    - B列填"计算天数"
    - 从 3 个透视表拉数据填到 C/D/G/L/M/O 列（不写公式，直接填值）
    """
    log = log_func or (lambda msg: None)
    log(f"[文件1] 读取: {file1_path}")

    # 文件1 也支持 csv 读取
    ext1 = os.path.splitext(file1_path)[1].lower()
    if ext1 == ".csv":
        # csv 输入：读到 A 列门店名后，新建成一个 xlsx 工作簿
        stores_csv = []
        for enc in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
            try:
                df1 = pd.read_csv(file1_path, header=None, encoding=enc).fillna("")
                stores_csv = [str(v).strip() for v in df1.iloc[:, 0].tolist() if str(v).strip()]
                break
            except UnicodeDecodeError:
                continue
        log(f"[文件1] CSV输入，门店数: {len(stores_csv)}")
        # 新建一个 xlsx 工作簿（csv 不支持 VLOOKUP 公式）
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "门店"
        for i, s in enumerate(stores_csv):
            ws.cell(row=i + 2, column=1, value=s)
    else:
        wb = load_workbook(file1_path)
        ws = wb.active
    log(f"[文件1] 工作表: {ws.title}")

    # 读A列门店名（跳过表头A1）
    stores = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is not None and str(v).strip():
            stores.append(str(v).strip())
    log(f"[文件1] A列门店数: {len(stores)}")

    # B列写计算天数
    for i in range(len(stores)):
        cell = ws.cell(row=i + 2, column=2)
        cell.value = days
        cell.font = Font(bold=False)

    # 按用户最新要求，VLOOKUP 结果填到指定列：
    #   C 列（推单数）              <- 文件2 P 列
    #   D 列（缺货导致退款订单数）  <- 文件2 R 列
    #   M 列（实付营业额）          <- 文件2 Q 列
    #   O 列（总时长）              <- 文件2 O 列
    # 文件2 透视结果 N列=门店, O=在线时长, P=推单数, Q=实付交易额, R=缺货退款订单数
    vlookup_map = [
        ("C", "P", "推单数"),
        ("D", "R", "缺货导致退款订单数"),
        ("M", "Q", "实付营业额"),
        ("O", "O", "总时长"),
    ]

    # 读透视表 xlsx，在内存里做 VLOOKUP（不写公式，直接填值）
    # 透视表布局：N列=门店, O=在线时长, P=推单数, Q=实付交易额, R=缺货退款订单数
    wb_pivot = load_workbook(file2_vlookup_path, data_only=False)
    ws_pivot = wb_pivot.active

    # 找 N 列（门店列）从哪一行开始有数据
    pivot_lookup = {}  # {门店名: (在线时长, 推单数, 实付交易额, 缺货退款订单数)}
    for row in ws_pivot.iter_rows(min_row=2, min_col=14, max_col=18, values_only=True):
        store_name, dur, orders, sales, refunds = row[0], row[1], row[2], row[3], row[4]
        if store_name is None or not str(store_name).strip():
            continue
        pivot_lookup[str(store_name).strip()] = (
            float(dur) if dur is not None else 0.0,
            float(orders) if orders is not None else 0.0,
            float(sales) if sales is not None else 0.0,
            float(refunds) if refunds is not None else 0.0,
        )
    log(f"[文件1] 透视表查找表构建完成: {len(pivot_lookup)} 家门店")

    # 按用户最新要求，直接填值到指定列：
    #   C 列（推单数）              <- 透视表 P 列
    #   D 列（缺货导致退款订单数）  <- 透视表 R 列
    #   M 列（实付营业额）          <- 透视表 Q 列
    #   O 列（总时长）              <- 透视表 O 列
    col_value_map = [
        ("C", 1, "推单数"),
        ("D", 3, "缺货导致退款订单数"),
        ("M", 2, "实付营业额"),
        ("O", 0, "总时长"),
    ]
    for i, store in enumerate(stores):
        r = i + 2
        matched = pivot_lookup.get(store, (0.0, 0.0, 0.0, 0.0))
        for col_out, val_idx, _ in col_value_map:
            ws[f"{col_out}{r}"] = matched[val_idx]

    # 读文件3透视表（双 sheet：中差评数 + 评分）
    if file3_vlookup_path:
        wb3 = load_workbook(file3_vlookup_path, data_only=False)
        # Sheet1: 中差评数 -> G 列
        ws3a = wb3["中差评数"] if "中差评数" in wb3.sheetnames else wb3.active
        lookup3_count = {}
        for row in ws3a.iter_rows(min_row=2, min_col=7, max_col=8, values_only=True):
            store_name, count = row[0], row[1]
            if store_name is None or not str(store_name).strip():
                continue
            try:
                lookup3_count[str(store_name).strip()] = int(count) if count is not None else 0
            except (ValueError, TypeError):
                lookup3_count[str(store_name).strip()] = 0
        log(f"[文件1] 文件3(中差评数)查找表: {len(lookup3_count)} 家门店")
        for i, store in enumerate(stores):
            r = i + 2
            ws[f"G{r}"] = lookup3_count.get(store, 0)
        # Sheet2: 评分 -> J 列
        if "评分" in wb3.sheetnames:
            ws3b = wb3["评分"]
            lookup3_score = {}
            for row in ws3b.iter_rows(min_row=2, min_col=7, max_col=8, values_only=True):
                store_name, score = row[0], row[1]
                if store_name is None or not str(store_name).strip():
                    continue
                try:
                    lookup3_score[str(store_name).strip()] = float(score) if score is not None else 0.0
                except (ValueError, TypeError):
                    lookup3_score[str(store_name).strip()] = 0.0
            log(f"[文件1] 文件3(评分)查找表: {len(lookup3_score)} 家门店")
            for i, store in enumerate(stores):
                r = i + 2
                ws[f"J{r}"] = lookup3_score.get(store, 0.0)
        else:
            log("[文件1] 未发现'评分' sheet，跳过 J 列填充")

    # 读文件4透视表（推广费），填到 L 列
    if file4_vlookup_path:
        wb4 = load_workbook(file4_vlookup_path, data_only=False)
        ws4 = wb4.active
        lookup4 = {}  # {门店: 推广费}
        for row in ws4.iter_rows(min_row=2, min_col=7, max_col=8, values_only=True):
            store_name, cost = row[0], row[1]
            if store_name is None or not str(store_name).strip():
                continue
            try:
                lookup4[str(store_name).strip()] = float(cost) if cost is not None else 0.0
            except (ValueError, TypeError):
                lookup4[str(store_name).strip()] = 0.0
        log(f"[文件1] 透视表4查找表: {len(lookup4)} 家门店")
        for i, store in enumerate(stores):
            r = i + 2
            ws[f"L{r}"] = lookup4.get(store, 0.0)

    # H 列：中差评率 = G / C（避免除零）
    for i in range(len(stores)):
        r = i + 2
        g_val = ws[f"G{r}"].value or 0
        c_val = ws[f"C{r}"].value or 0
        if isinstance(c_val, (int, float)) and c_val > 0:
            ws[f"H{r}"] = round(float(g_val) / float(c_val), 4)
        else:
            ws[f"H{r}"] = 0

    # R 列：商品(均分参考) — 每行填 product_ref
    # X 列：体验分 — 每行填 experience_score
    log(f"[文件1] 默认值: R(商品均分参考)={product_ref}, X(体验分)={experience_score}")
    for i in range(len(stores)):
        r = i + 2
        ws[f"R{r}"] = product_ref
        ws[f"X{r}"] = experience_score

    # 列宽
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    for c in ["C", "D", "E", "F", "G", "H", "L", "M", "O"]:
        ws.column_dimensions[c].width = 16

    wb.save(output_path)
    log(f"[文件1] 已保存: {output_path}")
    return output_path


# ============ GUI ============

class Worker(QThread):
    """后台处理线程"""
    log_signal = pyqtSignal(str)
    done_signal = pyqtSignal(bool, str)

    def __init__(self, file1, file2, days, output, file3=None, file4=None,
                 product_ref=0.9, experience_score=80.0):
        super().__init__()
        self.file1 = file1
        self.file2 = file2
        self.file3 = file3
        self.file4 = file4
        self.days = days
        self.output = output
        self.product_ref = product_ref
        self.experience_score = experience_score

    def log(self, msg):
        self.log_signal.emit(msg)

    def run(self):
        try:
            # 读A列门店名
            from openpyxl import load_workbook
            wb1 = load_workbook(self.file1)
            ws1 = wb1.active
            stores = []
            for row in ws1.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
                v = row[0]
                if v is not None and str(v).strip():
                    stores.append(str(v).strip())
            self.log(f"[文件1] A列门店数: {len(stores)}")

            tmp_dir = os.path.dirname(self.output)

            # 临时文件2处理结果
            tmp_file2 = os.path.join(tmp_dir, "_tmp_透视结果_文件2.xlsx")
            make_pivot_file2(self.file2, stores, tmp_file2, self.log)

            # 文件3（评价分析明细）
            tmp_file3 = None
            if self.file3 and os.path.exists(self.file3):
                tmp_file3 = os.path.join(tmp_dir, "_tmp_透视结果_文件3.xlsx")
                make_pivot_file3(self.file3, stores, tmp_file3, self.log)

            # 文件4（门店推广费）
            tmp_file4 = None
            if self.file4 and os.path.exists(self.file4):
                tmp_file4 = os.path.join(tmp_dir, "_tmp_透视结果_文件4.xlsx")
                make_pivot_file4(self.file4, stores, tmp_file4, self.log)

            # 处理文件1
            process_file1(self.file1, self.days, tmp_file2, self.output, self.log,
                          file3_vlookup_path=tmp_file3, file4_vlookup_path=tmp_file4,
                          product_ref=self.product_ref, experience_score=self.experience_score)

            self.done_signal.emit(True, f"处理完成！\n输出: {self.output}\n中间文件: {tmp_file2}{', ' + tmp_file3 if tmp_file3 else ''}{', ' + tmp_file4 if tmp_file4 else ''}")
        except Exception as e:
            import traceback
            self.log(f"错误: {traceback.format_exc()}")
            self.done_signal.emit(False, f"处理失败：{e}")


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("闪电仓计算工具 v1.0")
        self.setGeometry(200, 200, 800, 600)

        self.file1_path = ""
        self.file2_path = ""
        self.file3_path = ""
        self.file4_path = ""

        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout()

        # 标题
        title = QLabel("⚡ 闪电仓计算工具")
        title_font = QFont()
        title_font.setPointSize(16)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # 文件选择区
        file_group = QGroupBox("📁 选择文件")
        file_layout = QGridLayout()

        # 文件1
        file_layout.addWidget(QLabel("1. 闪电仓计算模板:"), 0, 0)
        self.file1_label = QLabel("(未选择)")
        self.file1_label.setStyleSheet("color: gray;")
        file_layout.addWidget(self.file1_label, 0, 1)
        btn1 = QPushButton("选择文件")
        btn1.clicked.connect(lambda: self._choose_file(1))
        file_layout.addWidget(btn1, 0, 2)

        # 文件2
        file_layout.addWidget(QLabel("2. 全店数据-门店成交明细:"), 1, 0)
        self.file2_label = QLabel("(未选择)")
        self.file2_label.setStyleSheet("color: gray;")
        file_layout.addWidget(self.file2_label, 1, 1)
        btn2 = QPushButton("选择文件")
        btn2.clicked.connect(lambda: self._choose_file(2))
        file_layout.addWidget(btn2, 1, 2)

        # 文件3：评价分析明细
        file_layout.addWidget(QLabel("3. 评价分析明细:"), 2, 0)
        self.file3_label = QLabel("(未选择)")
        self.file3_label.setStyleSheet("color: gray;")
        file_layout.addWidget(self.file3_label, 2, 1)
        btn3 = QPushButton("选择文件")
        btn3.clicked.connect(lambda: self._choose_file(3))
        file_layout.addWidget(btn3, 2, 2)

        # 文件4：门店推广费
        file_layout.addWidget(QLabel("4. 门店推广费:"), 3, 0)
        self.file4_label = QLabel("(未选择)")
        self.file4_label.setStyleSheet("color: gray;")
        file_layout.addWidget(self.file4_label, 3, 1)
        btn4 = QPushButton("选择文件")
        btn4.clicked.connect(lambda: self._choose_file(4))
        file_layout.addWidget(btn4, 3, 2)

        file_group.setLayout(file_layout)
        layout.addWidget(file_group)

        # 计算天数 + 默认值设置
        days_group = QGroupBox("⚙️ 设置")
        days_layout = QHBoxLayout()
        days_layout.addWidget(QLabel("计算天数:"))
        self.days_input = QLineEdit("30")
        self.days_input.setFixedWidth(80)
        days_layout.addWidget(self.days_input)
        days_layout.addSpacing(20)
        days_layout.addWidget(QLabel("商品(均分参考):"))
        self.product_ref_input = QLineEdit("0.9")
        self.product_ref_input.setFixedWidth(80)
        days_layout.addWidget(self.product_ref_input)
        days_layout.addSpacing(20)
        days_layout.addWidget(QLabel("体验分:"))
        self.experience_input = QLineEdit("80")
        self.experience_input.setFixedWidth(80)
        days_layout.addWidget(self.experience_input)
        days_layout.addStretch()
        days_group.setLayout(days_layout)
        layout.addWidget(days_group)

        # 开始按钮
        self.start_btn = QPushButton("🚀 开始处理")
        self.start_btn.setFixedHeight(40)
        self.start_btn.setStyleSheet("font-size: 14px; font-weight: bold; background: #4CAF50; color: white;")
        self.start_btn.clicked.connect(self._start)
        layout.addWidget(self.start_btn)

        # 日志区
        log_group = QGroupBox("📋 处理日志")
        log_layout = QVBoxLayout()
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setStyleSheet("background: #1e1e1e; color: #d4d4d4; font-family: Consolas, monospace;")
        log_layout.addWidget(self.log_text)
        log_group.setLayout(log_layout)
        layout.addWidget(log_group)

        self.setLayout(layout)

    def _choose_file(self, idx):
        # 支持所有常见表格后缀
        filter_str = (
            "所有表格文件 (*.xlsx *.xls *.xlsm *.csv);;"
            "Excel 文件 (*.xlsx *.xlsm);;"
            "Excel 97-2003 (*.xls);;"
            "CSV 文件 (*.csv);;"
            "所有文件 (*.*)"
        )
        path, _ = QFileDialog.getOpenFileName(
            self, f"选择文件 {idx}",
            "", filter_str
        )
        if path:
            if idx == 1:
                self.file1_path = path
                self.file1_label.setText(os.path.basename(path))
                self.file1_label.setStyleSheet("color: black;")
            elif idx == 2:
                self.file2_path = path
                self.file2_label.setText(os.path.basename(path))
                self.file2_label.setStyleSheet("color: black;")
            elif idx == 3:
                self.file3_path = path
                self.file3_label.setText(os.path.basename(path))
                self.file3_label.setStyleSheet("color: black;")
            elif idx == 4:
                self.file4_path = path
                self.file4_label.setText(os.path.basename(path))
                self.file4_label.setStyleSheet("color: black;")

    def _log(self, msg):
        self.log_text.append(msg)
        QApplication.processEvents()

    def _start(self):
        # 校验
        if not self.file1_path:
            QMessageBox.warning(self, "提示", "请先选择文件1（闪电仓计算模板）")
            return
        if not self.file2_path:
            QMessageBox.warning(self, "提示", "请先选择文件2（全店数据）")
            return
        try:
            days = int(self.days_input.text())
            product_ref = float(self.product_ref_input.text())
            experience_score = float(self.experience_input.text())
        except ValueError:
            QMessageBox.warning(self, "提示", "计算天数必须是整数、商品(均分参考)/体验分必须是数字")
            return

        # 选择输出路径（也支持多种格式）
        default_name = "闪电仓计算结果.xlsx"
        save_filter = (
            "Excel 文件 (*.xlsx);;"
            "Excel 97-2003 (*.xls);;"
            "CSV 文件 (*.csv)"
        )
        out, sel_filter = QFileDialog.getSaveFileName(
            self, "保存输出", default_name, save_filter
        )
        if not out:
            return

        # 提醒：VLOOKUP 公式需要 xlsx，csv 不支持
        if out.lower().endswith(".csv"):
            QMessageBox.warning(
                self, "输出格式提醒",
                "VLOOKUP 公式需要 Excel 格式才能保留\n请将输出文件后缀改为 .xlsx 或 .xls"
            )
            return

        # 启动后台线程
        self.start_btn.setEnabled(False)
        self.start_btn.setText("处理中...")
        self.worker = Worker(self.file1_path, self.file2_path, days, out, self.file3_path, self.file4_path,
                             product_ref=product_ref, experience_score=experience_score)
        self.worker.log_signal.connect(self._log)
        self.worker.done_signal.connect(self._on_done)
        self.worker.start()

    def _on_done(self, ok, msg):
        self.start_btn.setEnabled(True)
        self.start_btn.setText("🚀 开始处理")
        if ok:
            QMessageBox.information(self, "完成", msg)
        else:
            QMessageBox.critical(self, "失败", msg)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec_())
